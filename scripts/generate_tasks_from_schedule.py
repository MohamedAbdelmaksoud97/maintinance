from __future__ import annotations

import argparse
import json
import os
import re
import uuid
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

ROOT = Path(r"f:\lubrication\LUBRICTION PLAN 2026+2027")
SYSTEM_START_DATE = date(2026, 7, 15)
NAMESPACE = uuid.UUID("769f3fc0-779a-4fc9-a4d6-bf474416af50")

MARK_WORK_TYPE = {
    "I": "inspection",
    "C": "oil_change",
    "G": "greasing",
}


def stable_id(*parts: Any) -> str:
    return str(uuid.uuid5(NAMESPACE, "|".join("" if part is None else str(part) for part in parts)))


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def text(value: Any) -> str:
    value = clean(value)
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def norm(value: Any) -> str:
    value = text(value).casefold()
    value = re.sub(r"\s+", " ", value)
    return value


def as_date(value: Any) -> date | None:
    value = clean(value)
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def mark_code(value: Any) -> str | None:
    value = text(value).upper()
    if not value or value.startswith("#"):
        return None
    if value in MARK_WORK_TYPE:
        return value
    return None


def source_sheet_name(material_kind: str) -> str:
    return "3. مخطط التشحيم" if material_kind == "grease" else "3. المخطط الزمني"


def source_row_key(path: str, equipment_code: Any, part_description: Any, point_name: Any = None) -> tuple[str, str, str, str]:
    return (path, norm(equipment_code), norm(part_description), norm(point_name))


def load_plan_items(cur: psycopg2.extensions.cursor) -> dict[tuple[str, str, str, str], list[dict[str, Any]]]:
    cur.execute(
        """
        select
          api.id as annual_plan_item_id,
          ap.source_file,
          ap.material_kind,
          e.id as equipment_id,
          e.equipment_code,
          mp.id as maintenance_point_id,
          mp.point_name,
          mp.part_description,
          mp.execution_condition,
          mp.material_id,
          api.planned_quantity,
          api.planned_quantity_unit,
          api.original_values
        from annual_plan_items api
        join annual_plans ap on ap.id = api.annual_plan_id
        join maintenance_points mp on mp.id = api.maintenance_point_id
        join equipment e on e.id = mp.equipment_id
        """
    )
    lookup: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in cur.fetchall():
        (
            item_id,
            source_file,
            material_kind,
            equipment_id,
            equipment_code,
            point_id,
            point_name,
            part_description,
            execution_condition,
            material_id,
            quantity,
            quantity_unit,
            original_values,
        ) = row
        item = {
            "annual_plan_item_id": item_id,
            "source_file": source_file,
            "material_kind": material_kind,
            "equipment_id": equipment_id,
            "equipment_code": equipment_code,
            "maintenance_point_id": point_id,
            "point_name": point_name,
            "part_description": part_description,
            "execution_condition": execution_condition,
            "material_id": material_id,
            "planned_quantity": quantity,
            "planned_quantity_unit": quantity_unit,
            "original_values": original_values or {},
        }
        if material_kind == "grease":
            lookup[source_row_key(source_file, equipment_code, part_description, point_name)].append(item)
        else:
            lookup[source_row_key(source_file, equipment_code, part_description)].append(item)
    return lookup


def find_item(
    lookup: dict[tuple[str, str, str, str], list[dict[str, Any]]],
    path: Path,
    material_kind: str,
    equipment_code: Any,
    part_description: Any,
    point_name: Any,
) -> dict[str, Any] | None:
    if material_kind == "grease":
        keys = [
            source_row_key(str(path), equipment_code, part_description, point_name),
            source_row_key(str(path), equipment_code, part_description),
        ]
    else:
        keys = [source_row_key(str(path), equipment_code, part_description)]

    for key in keys:
        matches = lookup.get(key)
        if matches:
            return matches[0]
    return None


def discover_plan_files() -> list[Path]:
    return sorted(path for path in ROOT.glob("* AREA/*.xlsx") if re.search(r"20(26|27)", path.name))


def collect_schedule_tasks(
    cur: psycopg2.extensions.cursor,
    start_date: date,
) -> tuple[list[tuple[Any, ...]], list[dict[str, Any]], dict[str, int]]:
    cur.execute("select code, id from maintenance_work_types where code in ('inspection', 'oil_change', 'greasing')")
    work_type_ids = dict(cur.fetchall())
    cur.execute("select id from task_statuses where code = 'NEEDS_ASSIGNMENT'")
    status_id = cur.fetchone()[0]
    cur.execute("select id from assignment_statuses where code = 'UNASSIGNED'")
    assignment_status_id = cur.fetchone()[0]

    lookup = load_plan_items(cur)
    tasks: list[tuple[Any, ...]] = []
    unmatched: list[dict[str, Any]] = []
    stats = {"files": 0, "marked_cells": 0, "matched": 0, "unmatched": 0}

    for path in discover_plan_files():
        material_kind = "grease" if "GREASE" in path.name.upper() else "oil"
        sheet_name = source_sheet_name(material_kind)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            continue

        stats["files"] += 1
        sheet = workbook[sheet_name]
        header = list(next(sheet.iter_rows(min_row=7, max_row=7, values_only=True)))
        date_columns = [(index, as_date(value)) for index, value in enumerate(header)]
        date_columns = [(index, value) for index, value in date_columns if value and value >= start_date]

        for row_number, row in enumerate(sheet.iter_rows(min_row=8, values_only=True), start=8):
            equipment_code = clean(row[1] if len(row) > 1 else None)
            if not equipment_code or str(equipment_code).casefold() == "equipment code":
                continue

            part_description = clean(row[3] if len(row) > 3 else None)
            point_name = clean(row[4] if material_kind == "grease" and len(row) > 4 else None)
            item = find_item(lookup, path, material_kind, equipment_code, part_description, point_name)

            for column_index, scheduled_date in date_columns:
                if column_index >= len(row):
                    continue
                mark = mark_code(row[column_index])
                if not mark:
                    continue

                stats["marked_cells"] += 1
                if not item:
                    stats["unmatched"] += 1
                    unmatched.append(
                        {
                            "source_file": str(path),
                            "source_sheet": sheet_name,
                            "source_row": row_number,
                            "source_column": column_index + 1,
                            "scheduled_date": scheduled_date.isoformat(),
                            "mark": mark,
                            "equipment_code": text(equipment_code),
                            "part_description": text(part_description),
                            "point_name": text(point_name),
                        }
                    )
                    continue

                work_type_code = MARK_WORK_TYPE[mark]
                quantity = item["planned_quantity"]
                quantity_unit = item["planned_quantity_unit"]
                if mark == "I":
                    quantity = None
                    quantity_unit = None

                original_values = dict(item["original_values"] or {})
                original_values.update(
                    {
                        "source_mode": "schedule_sheet",
                        "source_file": str(path),
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                        "source_column": column_index + 1,
                        "scheduled_date": scheduled_date.isoformat(),
                        "schedule_mark": mark,
                    }
                )

                task_id = stable_id("schedule-task", str(path), sheet_name, row_number, column_index + 1, scheduled_date.isoformat(), mark)
                tasks.append(
                    (
                        task_id,
                        item["annual_plan_item_id"],
                        item["equipment_id"],
                        item["maintenance_point_id"],
                        work_type_ids[work_type_code],
                        item["material_id"],
                        status_id,
                        assignment_status_id,
                        scheduled_date,
                        scheduled_date,
                        item["execution_condition"],
                        quantity,
                        quantity_unit,
                        psycopg2.extras.Json(original_values),
                    )
                )
                stats["matched"] += 1

        workbook.close()

    return tasks, unmatched, stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate planned tasks directly from daily schedule sheets.")
    parser.add_argument("--start-date", default=SYSTEM_START_DATE.isoformat())
    parser.add_argument("--apply", action="store_true", help="Write changes to the database.")
    args = parser.parse_args()

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL is required")

    start_date = datetime.strptime(args.start_date, "%Y-%m-%d").date()
    with psycopg2.connect(database_url, sslmode="require") as conn:
        with conn.cursor() as cur:
            tasks, unmatched, stats = collect_schedule_tasks(cur, start_date)

            if args.apply:
                cur.execute("select id from task_statuses where code = 'OLD'")
                old_status_id = cur.fetchone()[0]
                cur.execute(
                    """
                    update planned_tasks
                    set status_id = %s,
                        updated_at = now(),
                        original_values = original_values || '{"replaced_by_schedule_sheet": true}'::jsonb
                    where scheduled_date >= %s
                      and not (original_values ? 'source_mode' and original_values->>'source_mode' = 'schedule_sheet')
                    """,
                    (old_status_id, start_date),
                )

                psycopg2.extras.execute_values(
                    cur,
                    """
                    insert into planned_tasks (
                      id, annual_plan_item_id, equipment_id, maintenance_point_id,
                      work_type_id, material_id, status_id, assignment_status_id,
                      original_due_date, scheduled_date, execution_condition,
                      planned_quantity, planned_quantity_unit, original_values
                    )
                    values %s
                    on conflict (id) do update set
                      annual_plan_item_id = excluded.annual_plan_item_id,
                      equipment_id = excluded.equipment_id,
                      maintenance_point_id = excluded.maintenance_point_id,
                      work_type_id = excluded.work_type_id,
                      material_id = excluded.material_id,
                      original_due_date = excluded.original_due_date,
                      scheduled_date = excluded.scheduled_date,
                      execution_condition = excluded.execution_condition,
                      planned_quantity = excluded.planned_quantity,
                      planned_quantity_unit = excluded.planned_quantity_unit,
                      original_values = excluded.original_values,
                      updated_at = now()
                    """,
                    tasks,
                    page_size=1000,
                )
                conn.commit()

    print(json.dumps({"stats": stats, "tasks": len(tasks), "unmatched": unmatched[:25]}, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
