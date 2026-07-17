from __future__ import annotations

import argparse
import json
import math
import os
import re
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from openpyxl.utils.datetime import from_excel
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb

ROOT = Path("plan files")
SYSTEM_START_DATE = date(2026, 7, 15)
GREASE_CHANGE_LAST_DATE = date(2025, 3, 30)
GREASE_CHANGE_FREQUENCY_HOURS = 20000
NAMESPACE = uuid.UUID("ba45c4a7-14a7-4c2e-a7c2-84f3cf15cb0b")
CONTINUOUS_OPERATION_AREAS = {"FINISH_MILL", "PACKHOUSE"}


@dataclass(frozen=True)
class ShutdownWindow:
    line_code: str
    starts_on: date
    ends_on: date
    source_file: str
    source_sheet: str
    plan_year: int | None
    original_values: dict[str, Any]


@dataclass(frozen=True)
class Operation:
    area_code: str
    area_name: str
    plan_year: int
    material_kind: str
    source_file: str
    source_sheet: str
    source_row: int
    equipment_code: str
    equipment_name: str | None
    line_code: str | None
    part_description: str | None
    point_name: str | None
    work_type_code: str
    execution_condition: str
    material_name: str | None
    quantity: float | None
    quantity_unit: str | None
    running_hours_per_day: float | None
    frequency_hours: float | None
    frequency_days: float | None
    last_date: date | None
    previous_raw_due_date: date | None
    previous_scheduled_date: date | None
    calendar_due_date: date | None
    raw_due_date: date | None
    scheduled_date: date | None
    needs_shutdown_date: bool
    shifted_by_shutdown_discount: bool
    data_quality_status: str
    original_values: dict[str, Any]


def stable_id(*parts: Any) -> str:
    return str(uuid.uuid5(NAMESPACE, "|".join("" if part is None else str(part) for part in parts)))


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def text(value: Any) -> str | None:
    value = clean(value)
    return str(value) if value is not None else None


def number(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        if isinstance(value, str) and not re.search(r"\d", value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def as_date(value: Any) -> date | None:
    value = clean(value)
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def normalize_code(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def normalize_line(value: Any) -> str | None:
    value = text(value)
    if not value:
        return None
    match = re.search(r"[12]", value)
    return match.group(0) if match else value


def area_from_file(path: Path) -> tuple[str, str]:
    name = path.name.upper()
    if "FINISH MILL" in name:
        return "FINISH_MILL", "FINISH MILL"
    if "PACKHOUSE" in name:
        return "PACKHOUSE", "PACKHOUSE"
    if "CRUSHER" in name:
        return "CRUSHER", "CRUSHER"
    if "KLIN" in name:
        return "KLIN", "KLIN"
    return "GENERAL", "GENERAL"


def year_from_file(path: Path) -> int | None:
    match = re.search(r"(20\d{2})", path.name)
    return int(match.group(1)) if match else None


def material_kind_from_file(path: Path) -> str:
    return "grease" if "GREASE" in path.name.upper() else "oil"


def ignores_shutdown_windows(area_code: str) -> bool:
    return area_code in CONTINUOUS_OPERATION_AREAS


def input_sheet(workbook: openpyxl.Workbook, material_kind: str) -> Any:
    starts = "2. مدخلات التشحيم" if material_kind == "grease" else "2. مدخلات المعدات"
    for sheet in workbook.worksheets:
        if sheet.title.startswith(starts):
            return sheet
    raise ValueError(f"Input sheet not found for {material_kind}")


def shutdown_sheet(workbook: openpyxl.Workbook) -> Any | None:
    for sheet in workbook.worksheets:
        if "توقيتات التوقف" in sheet.title:
            return sheet
    return None


def step_days_from_days(frequency_days: float | None) -> int | None:
    if not frequency_days or frequency_days <= 0:
        return None
    return max(1, math.floor(float(frequency_days)))


def step_days_from_hours(frequency_hours: float | None, running_hours: float | None) -> int | None:
    if not frequency_hours or not running_hours or running_hours <= 0:
        return None
    return max(1, math.floor(float(frequency_hours) / float(running_hours)))


def first_next_due(last_date: date | None, step_days: int | None, as_of: date) -> date | None:
    if not last_date or not step_days:
        return None
    current = last_date + timedelta(days=step_days)
    while current < as_of:
        current += timedelta(days=step_days)
    return current


def in_window(day: date, window: ShutdownWindow) -> bool:
    return window.starts_on <= day <= window.ends_on


def is_shutdown_day(day: date, line_code: str | None, windows_by_line: dict[str, list[ShutdownWindow]]) -> bool:
    if not line_code:
        return False
    return any(in_window(day, window) for window in windows_by_line.get(line_code, []))


def add_operating_days(
    anchor: date | None,
    operating_days: int | None,
    line_code: str | None,
    windows_by_line: dict[str, list[ShutdownWindow]],
) -> date | None:
    if not anchor or operating_days is None:
        return None
    if not line_code:
        return anchor + timedelta(days=max(0, operating_days))

    current = anchor
    counted = 0
    while counted < operating_days:
        current += timedelta(days=1)
        if not is_shutdown_day(current, line_code, windows_by_line):
            counted += 1
    return current


def subtract_operating_days(
    scheduled: date | None,
    operating_days: int | None,
    line_code: str | None,
    windows_by_line: dict[str, list[ShutdownWindow]],
) -> date | None:
    if not scheduled or operating_days is None:
        return None
    if not line_code:
        return scheduled - timedelta(days=max(0, operating_days))

    current = scheduled
    counted = 0
    while counted < operating_days:
        current -= timedelta(days=1)
        if not is_shutdown_day(current, line_code, windows_by_line):
            counted += 1
    return current


def first_next_operating_due(
    last_date: date | None,
    step_days: int | None,
    as_of: date,
    line_code: str | None,
    windows_by_line: dict[str, list[ShutdownWindow]],
) -> date | None:
    if not last_date or not step_days:
        return None
    current = add_operating_days(last_date, step_days, line_code, windows_by_line)
    while current and current < as_of:
        current = add_operating_days(current, step_days, line_code, windows_by_line)
    return current


def adjust_for_shutdown(
    raw_due: date | None,
    line_code: str | None,
    execution_condition: str,
    windows_by_line: dict[str, list[ShutdownWindow]],
) -> tuple[date | None, bool]:
    if not raw_due or not line_code:
        return raw_due, execution_condition == "shutdown"
    windows = windows_by_line.get(line_code, [])
    if not windows:
        return raw_due, execution_condition == "shutdown"

    if execution_condition == "shutdown":
        for window in windows:
            if in_window(raw_due, window):
                return raw_due, False
            if raw_due < window.starts_on:
                return window.starts_on, False
        return raw_due, True

    scheduled = raw_due
    moved = True
    while moved:
        moved = False
        for window in windows:
            if in_window(scheduled, window):
                scheduled = window.ends_on + timedelta(days=1)
                moved = True
                break
    return scheduled, False


def collect_shutdown_windows(paths: list[Path]) -> list[ShutdownWindow]:
    windows: dict[tuple[str, date, date], ShutdownWindow] = {}
    for path in paths:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = shutdown_sheet(workbook)
        if not sheet:
            workbook.close()
            continue
        plan_year = year_from_file(path)
        line_specs = [("1", 1, 2, 3), ("2", 5, 6, 7)]
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            for line_code, start_col, end_col, desc_col in line_specs:
                starts_on = as_date(row[start_col - 1] if len(row) >= start_col else None)
                ends_on = as_date(row[end_col - 1] if len(row) >= end_col else None)
                if not starts_on or not ends_on:
                    continue
                key = (line_code, starts_on, ends_on)
                windows[key] = ShutdownWindow(
                    line_code=line_code,
                    starts_on=starts_on,
                    ends_on=ends_on,
                    source_file=str(path),
                    source_sheet=sheet.title,
                    plan_year=plan_year,
                    original_values={"source_row": row_number, "line_code": line_code},
                )
        workbook.close()
    return sorted(windows.values(), key=lambda item: (item.line_code, item.starts_on, item.ends_on))


def build_operation(
    *,
    path: Path,
    sheet_title: str,
    row_number: int,
    row: tuple[Any, ...],
    work_type_code: str,
    execution_condition: str,
    material_kind: str,
    as_of: date,
    windows_by_line: dict[str, list[ShutdownWindow]],
) -> Operation | None:
    area_code, area_name = area_from_file(path)
    plan_year = year_from_file(path)
    if not plan_year:
        return None

    equipment_code = text(row[0] if len(row) > 0 else None)
    if not equipment_code or equipment_code.casefold() in {"كود المعدة", "equipment code"}:
        return None

    equipment_name = text(row[1] if len(row) > 1 else None)
    part_description = text(row[2] if len(row) > 2 else None)
    line_code = normalize_line(row[3] if len(row) > 3 else None)
    schedule_line_code = None if ignores_shutdown_windows(area_code) else line_code
    running_hours = number(row[6] if len(row) > 6 else None)

    if material_kind == "oil":
        material_name = text(row[4] if len(row) > 4 else None)
        quantity = number(row[5] if len(row) > 5 else None)
        point_name = part_description
        quantity_unit = "L"
        if work_type_code == "inspection":
            frequency_days = number(row[8] if len(row) > 8 else None)
            frequency_hours = None
            last_date = as_date(row[10] if len(row) > 10 else None)
            step_days = step_days_from_days(frequency_days)
            material_for_task = None
            quantity_for_task = None
            quantity_unit_for_task = None
        else:
            frequency_hours = number(row[7] if len(row) > 7 else None)
            frequency_days = None
            last_date = as_date(row[9] if len(row) > 9 else None)
            step_days = step_days_from_hours(frequency_hours, running_hours)
            material_for_task = material_name
            quantity_for_task = quantity
            quantity_unit_for_task = quantity_unit
    else:
        point_name = text(row[4] if len(row) > 4 else None)
        material_name = text(row[5] if len(row) > 5 else None)
        quantity = number(row[8] if len(row) > 8 else None)
        quantity_unit = "g"
        if work_type_code == "greasing":
            frequency_days = number(row[10] if len(row) > 10 else None)
            frequency_hours = None
            last_date = as_date(row[11] if len(row) > 11 else None)
            step_days = step_days_from_days(frequency_days)
        else:
            frequency_days = None
            frequency_hours = GREASE_CHANGE_FREQUENCY_HOURS
            last_date = GREASE_CHANGE_LAST_DATE
            step_days = step_days_from_hours(frequency_hours, running_hours)
        material_for_task = material_name
        quantity_for_task = quantity
        quantity_unit_for_task = quantity_unit

    uses_operating_days = work_type_code in {"inspection", "greasing"}
    calendar_due = first_next_due(last_date, step_days, as_of)
    raw_due = (
        first_next_operating_due(last_date, step_days, as_of, schedule_line_code, windows_by_line)
        if uses_operating_days
        else calendar_due
    )
    if ignores_shutdown_windows(area_code):
        scheduled_date, needs_shutdown_date = raw_due, False
    else:
        scheduled_date, needs_shutdown_date = adjust_for_shutdown(raw_due, schedule_line_code, execution_condition, windows_by_line)
    previous_raw_due = (
        subtract_operating_days(raw_due, step_days, schedule_line_code, windows_by_line)
        if uses_operating_days
        else (raw_due - timedelta(days=step_days) if raw_due and step_days else None)
    )
    if ignores_shutdown_windows(area_code):
        previous_scheduled_date = previous_raw_due
    else:
        previous_scheduled_date, _ = adjust_for_shutdown(previous_raw_due, schedule_line_code, execution_condition, windows_by_line)
    if uses_operating_days:
        previous_scheduled_date = previous_raw_due
    shifted_by_shutdown_discount = bool(uses_operating_days and raw_due and calendar_due and raw_due != calendar_due)

    has_required_timing = bool(scheduled_date and step_days)
    needs_line_review = line_code is None and not ignores_shutdown_windows(area_code)
    status = "COMPLETE" if has_required_timing and not needs_line_review else "MISSING_DATA"
    if work_type_code in {"inspection", "greasing"} and not last_date:
        status = "MISSING_DATA"

    original_values = {
        "source_mode": "calculated_next_due",
        "source_file": str(path),
        "source_sheet": sheet_title,
        "source_row": row_number,
        "line_code": line_code,
        "schedule_line_code": schedule_line_code,
        "ignores_shutdown_windows": ignores_shutdown_windows(area_code),
        "last_date": last_date.isoformat() if last_date else None,
        "frequency_hours": frequency_hours,
        "frequency_days": frequency_days,
        "running_hours_per_day": running_hours,
        "step_days": step_days,
        "uses_operating_days": uses_operating_days,
        "previous_raw_due_date": previous_raw_due.isoformat() if previous_raw_due else None,
        "previous_scheduled_date": previous_scheduled_date.isoformat() if previous_scheduled_date else None,
        "calendar_due_without_shutdown_discount": calendar_due.isoformat() if calendar_due else None,
        "shifted_by_shutdown_discount": shifted_by_shutdown_discount,
        "raw_due_date": raw_due.isoformat() if raw_due else None,
        "scheduled_date": scheduled_date.isoformat() if scheduled_date else None,
        "needs_shutdown_date": needs_shutdown_date,
    }

    return Operation(
        area_code=area_code,
        area_name=area_name,
        plan_year=plan_year,
        material_kind=material_kind,
        source_file=str(path),
        source_sheet=sheet_title,
        source_row=row_number,
        equipment_code=equipment_code,
        equipment_name=equipment_name,
        line_code=line_code,
        part_description=part_description,
        point_name=point_name,
        work_type_code=work_type_code,
        execution_condition=execution_condition,
        material_name=material_for_task,
        quantity=quantity_for_task,
        quantity_unit=quantity_unit_for_task,
        running_hours_per_day=running_hours,
        frequency_hours=frequency_hours,
        frequency_days=frequency_days,
        last_date=last_date,
        previous_raw_due_date=previous_raw_due,
        previous_scheduled_date=previous_scheduled_date,
        calendar_due_date=calendar_due,
        raw_due_date=raw_due,
        scheduled_date=scheduled_date,
        needs_shutdown_date=needs_shutdown_date,
        shifted_by_shutdown_discount=shifted_by_shutdown_discount,
        data_quality_status=status,
        original_values=original_values,
    )


def collect_operations(paths: list[Path], as_of: date, windows: list[ShutdownWindow]) -> list[Operation]:
    windows_by_line: dict[str, list[ShutdownWindow]] = defaultdict(list)
    for window in windows:
        windows_by_line[window.line_code].append(window)
    for line_windows in windows_by_line.values():
        line_windows.sort(key=lambda item: item.starts_on)

    operations: list[Operation] = []
    for path in paths:
        if "Master" in path.name or path.name.startswith("~$"):
            continue
        material_kind = material_kind_from_file(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = input_sheet(workbook, material_kind)
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            if not any(value is not None for value in row):
                continue
            if material_kind == "oil":
                specs = [("inspection", "running"), ("oil_change", "shutdown")]
            else:
                specs = [("greasing", "running"), ("grease_change", "shutdown")]
            for work_type_code, execution_condition in specs:
                operation = build_operation(
                    path=path,
                    sheet_title=sheet.title,
                    row_number=row_number,
                    row=row,
                    work_type_code=work_type_code,
                    execution_condition=execution_condition,
                    material_kind=material_kind,
                    as_of=as_of,
                    windows_by_line=windows_by_line,
                )
                if operation:
                    operations.append(operation)
        workbook.close()
    return operations


def load_equipment_lookup(conn: psycopg.Connection) -> dict[str, str]:
    rows = conn.execute("select id, equipment_code, is_active, original_values from equipment").fetchall()
    preferred: dict[str, str] = {}
    fallback: dict[str, str] = {}
    for row in rows:
        key = normalize_code(row["equipment_code"])
        fallback.setdefault(key, row["id"])
        values = row["original_values"] or {}
        if row["is_active"] and values.get("source_mode") in {"master_equipment", "manual_equipment"}:
            preferred[key] = row["id"]
    return {**fallback, **preferred}


def material_code(kind: str, name: str) -> str:
    return f"{kind}-{uuid.uuid5(NAMESPACE, name).hex[:10]}"


def apply_to_database(database_url: str, operations: list[Operation], windows: list[ShutdownWindow], as_of: date) -> dict[str, int]:
    with psycopg.connect(database_url, row_factory=dict_row) as conn:
        def execute_many(sql: str, rows: list[tuple[Any, ...]]) -> None:
            if not rows:
                return
            with conn.cursor() as cur:
                cur.executemany(sql, rows)

        work_type_rows = conn.execute("select code, id from maintenance_work_types").fetchall()
        work_type_ids = {row["code"]: row["id"] for row in work_type_rows}
        area_lookup = {row["code"]: row["id"] for row in conn.execute("select id, code from areas").fetchall()}
        line_lookup = {
            (row["area_id"], row["line_code"]): row["id"]
            for row in conn.execute("select id, area_id, line_code from production_lines").fetchall()
        }
        status_id = conn.execute("select id from task_statuses where code = 'NEEDS_ASSIGNMENT'").fetchone()["id"]
        old_status_id = conn.execute("select id from task_statuses where code = 'OLD'").fetchone()["id"]
        assignment_status_id = conn.execute("select id from assignment_statuses where code = 'UNASSIGNED'").fetchone()["id"]
        equipment_lookup = load_equipment_lookup(conn)

        area_rows = {}
        line_rows = {}
        material_rows = {}
        equipment_rows = {}
        plan_rows = {}
        point_rows = []
        item_rows = []
        task_rows = []
        point_ids: list[str] = []
        task_ids: list[str] = []
        shutdown_rows = []

        for window in windows:
            shutdown_rows.append(
                (
                    stable_id("shutdown", window.line_code, window.starts_on, window.ends_on),
                    window.line_code,
                    window.starts_on,
                    window.ends_on,
                    window.source_file,
                    window.source_sheet,
                    window.plan_year,
                    Jsonb(window.original_values),
                )
            )

        for op in operations:
            area_id = area_lookup.get(op.area_code, stable_id("area", op.area_code))
            area_rows[area_id] = (area_id, op.area_code, op.area_name)
            line_id = None
            if op.line_code:
                line_id = line_lookup.get((area_id, op.line_code), stable_id("line", op.area_code, op.line_code))
                line_rows[line_id] = (line_id, area_id, op.line_code, f"Line {op.line_code}")

            equipment_id = equipment_lookup.get(normalize_code(op.equipment_code))
            if not equipment_id:
                equipment_id = stable_id("equipment", op.area_code, op.equipment_code)
                equipment_rows[equipment_id] = (
                    equipment_id,
                    area_id,
                    line_id,
                    op.equipment_code,
                    op.equipment_name,
                    op.part_description,
                    Jsonb({"source_mode": "calculated_plan_equipment", "line_code": op.line_code}),
                    op.data_quality_status,
                )

            material_id = None
            if op.material_name:
                material_id = stable_id("material", op.material_kind, op.material_name)
                material_rows[material_id] = (
                    material_id,
                    op.material_kind,
                    material_code(op.material_kind, op.material_name),
                    op.material_name,
                    op.quantity_unit,
                    "COMPLETE",
                )

            plan_id = stable_id("annual_plan", op.area_code, op.plan_year, op.material_kind, Path(op.source_file).name)
            plan_rows[plan_id] = (plan_id, area_id, op.plan_year, op.material_kind, op.source_file)
            point_id = stable_id("point", op.source_file, op.source_row, op.work_type_code)
            item_id = stable_id("plan_item", op.source_file, op.source_row, op.work_type_code)
            original_json = Jsonb(op.original_values)
            point_ids.append(point_id)
            point_rows.append(
                (
                    point_id,
                    equipment_id,
                    work_type_ids[op.work_type_code],
                    material_id,
                    op.point_name,
                    op.part_description,
                    op.execution_condition,
                    op.quantity,
                    op.quantity_unit,
                    op.running_hours_per_day,
                    op.frequency_hours,
                    op.frequency_days,
                    op.last_date if op.work_type_code in {"oil_change", "grease_change"} else None,
                    op.last_date if op.work_type_code == "inspection" else None,
                    op.last_date if op.work_type_code == "greasing" else None,
                    original_json,
                    op.data_quality_status,
                )
            )
            item_rows.append(
                (
                    item_id,
                    plan_id,
                    point_id,
                    op.quantity,
                    op.quantity_unit,
                    op.frequency_hours,
                    op.frequency_days,
                    original_json,
                    op.data_quality_status,
                )
            )
            if op.scheduled_date:
                task_id = stable_id("task", item_id, op.scheduled_date.isoformat(), op.work_type_code)
                task_ids.append(task_id)
                task_rows.append(
                    (
                        task_id,
                        item_id,
                        equipment_id,
                        point_id,
                        work_type_ids[op.work_type_code],
                        material_id,
                        status_id,
                        assignment_status_id,
                        op.raw_due_date or op.scheduled_date,
                        op.scheduled_date,
                        op.execution_condition,
                        op.quantity,
                        op.quantity_unit,
                        original_json,
                    )
                )

        with conn.transaction():
            execute_many(
                """
                insert into shutdown_windows (id,line_code,starts_on,ends_on,source_file,source_sheet,plan_year,original_values)
                values (%s,%s,%s,%s,%s,%s,%s,%s)
                on conflict (line_code, starts_on, ends_on) do update set
                  source_file = excluded.source_file,
                  source_sheet = excluded.source_sheet,
                  plan_year = excluded.plan_year,
                  original_values = excluded.original_values,
                  is_active = true,
                  updated_at = now()
                """,
                shutdown_rows,
            )
            execute_many(
                "insert into areas (id,code,name) values (%s,%s,%s) on conflict (code) do update set name=excluded.name",
                list(area_rows.values()),
            )
            execute_many(
                """
                insert into production_lines (id,area_id,line_code,name)
                values (%s,%s,%s,%s)
                on conflict (area_id, line_code) do update set name=excluded.name
                """,
                list(line_rows.values()),
            )
            execute_many(
                """
                insert into materials (id,material_kind,code,name,unit,data_quality_status)
                values (%s,%s,%s,%s,%s,%s)
                on conflict (id) do update set name=excluded.name, unit=excluded.unit, is_active=true
                """,
                list(material_rows.values()),
            )
            execute_many(
                """
                insert into equipment (id,area_id,production_line_id,equipment_code,name,description,original_values,data_quality_status)
                values (%s,%s,%s,%s,%s,%s,%s,%s)
                on conflict (id) do update set
                  production_line_id=excluded.production_line_id,
                  name=coalesce(excluded.name, equipment.name),
                  description=coalesce(excluded.description, equipment.description),
                  updated_at=now()
                """,
                list(equipment_rows.values()),
            )
            execute_many(
                """
                insert into annual_plans (id,area_id,plan_year,material_kind,source_file)
                values (%s,%s,%s,%s,%s)
                on conflict (id) do update set updated_at=now()
                """,
                list(plan_rows.values()),
            )
            execute_many(
                """
                insert into maintenance_points (
                  id,equipment_id,work_type_id,material_id,point_name,part_description,
                  execution_condition,quantity,quantity_unit,running_hours_per_day,
                  frequency_hours,frequency_days,last_change_date,last_inspection_date,last_grease_date,
                  original_values,data_quality_status
                )
                values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                on conflict (id) do update set
                  material_id=excluded.material_id,
                  point_name=excluded.point_name,
                  part_description=excluded.part_description,
                  execution_condition=excluded.execution_condition,
                  quantity=excluded.quantity,
                  quantity_unit=excluded.quantity_unit,
                  running_hours_per_day=excluded.running_hours_per_day,
                  frequency_hours=excluded.frequency_hours,
                  frequency_days=excluded.frequency_days,
                  last_change_date=excluded.last_change_date,
                  last_inspection_date=excluded.last_inspection_date,
                  last_grease_date=excluded.last_grease_date,
                  original_values=excluded.original_values,
                  data_quality_status=excluded.data_quality_status,
                  is_active=true,
                  updated_at=now()
                """,
                point_rows,
            )
            execute_many(
                """
                insert into annual_plan_items (
                  id,annual_plan_id,maintenance_point_id,planned_quantity,planned_quantity_unit,
                  frequency_hours,frequency_days,original_values,data_quality_status
                )
                values (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                on conflict (id) do update set
                  planned_quantity=excluded.planned_quantity,
                  planned_quantity_unit=excluded.planned_quantity_unit,
                  frequency_hours=excluded.frequency_hours,
                  frequency_days=excluded.frequency_days,
                  original_values=excluded.original_values,
                  data_quality_status=excluded.data_quality_status,
                  updated_at=now()
                """,
                item_rows,
            )
            conn.execute(
                """
                update planned_tasks
                set status_id = %s,
                    updated_at = now(),
                    original_values = original_values || '{"replaced_by_calculated_plan": true}'::jsonb
                where scheduled_date >= %s
                  and not (original_values ? 'source_mode' and original_values->>'source_mode' = 'calculated_next_due')
                  and completed_at is null
                  and not exists (select 1 from execution_reports er where er.task_id = planned_tasks.id)
                  and not exists (select 1 from non_execution_reports ner where ner.task_id = planned_tasks.id)
                """,
                (old_status_id, SYSTEM_START_DATE),
            )
            execute_many(
                """
                insert into planned_tasks (
                  id,annual_plan_item_id,equipment_id,maintenance_point_id,work_type_id,material_id,
                  status_id,assignment_status_id,original_due_date,scheduled_date,execution_condition,
                  planned_quantity,planned_quantity_unit,original_values
                )
                values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                on conflict (id) do update set
                  annual_plan_item_id=excluded.annual_plan_item_id,
                  equipment_id=excluded.equipment_id,
                  maintenance_point_id=excluded.maintenance_point_id,
                  work_type_id=excluded.work_type_id,
                  material_id=excluded.material_id,
                  original_due_date=excluded.original_due_date,
                  scheduled_date=excluded.scheduled_date,
                  execution_condition=excluded.execution_condition,
                  planned_quantity=excluded.planned_quantity,
                  planned_quantity_unit=excluded.planned_quantity_unit,
                  original_values=excluded.original_values,
                  updated_at=now()
                """,
                task_rows,
            )
            if point_ids:
                conn.execute(
                    """
                    update planned_tasks
                    set status_id = %s,
                        updated_at = now(),
                        original_values = original_values || '{"replaced_by_operating_days_import": true}'::jsonb
                    where maintenance_point_id = any(%s::uuid[])
                      and not (id = any(%s::uuid[]))
                      and scheduled_date >= %s
                      and completed_at is null
                      and original_values->>'source_mode' in ('calculated_next_due', 'dynamic_plan')
                      and not exists (select 1 from execution_reports er where er.task_id = planned_tasks.id)
                      and not exists (select 1 from non_execution_reports ner where ner.task_id = planned_tasks.id)
                    """,
                    (old_status_id, point_ids, task_ids or ["00000000-0000-0000-0000-000000000000"], as_of),
                )
            conn.execute(
                """
                update planned_tasks
                set status_id = %s,
                    updated_at = now(),
                    original_values = planned_tasks.original_values || '{"replaced_by_continuous_area_rule": true}'::jsonb
                from equipment
                join areas on areas.id = equipment.area_id
                where planned_tasks.equipment_id = equipment.id
                  and areas.code = any(%s::text[])
                  and coalesce(planned_tasks.original_values->>'ignores_shutdown_windows', 'false') <> 'true'
                  and planned_tasks.scheduled_date >= %s
                  and planned_tasks.completed_at is null
                  and planned_tasks.original_values->>'source_mode' in ('calculated_next_due', 'dynamic_plan')
                  and not exists (select 1 from execution_reports er where er.task_id = planned_tasks.id)
                  and not exists (select 1 from non_execution_reports ner where ner.task_id = planned_tasks.id)
                """,
                (old_status_id, sorted(CONTINUOUS_OPERATION_AREAS), as_of),
            )

        return {
            "shutdown_windows": len(shutdown_rows),
            "areas": len(area_rows),
            "production_lines": len(line_rows),
            "materials": len(material_rows),
            "fallback_equipment_inserted": len(equipment_rows),
            "maintenance_points": len(point_rows),
            "annual_plan_items": len(item_rows),
            "planned_tasks": len(task_rows),
        }


def summarize(operations: list[Operation], windows: list[ShutdownWindow]) -> dict[str, Any]:
    by_type = defaultdict(int)
    complete = 0
    missing = 0
    needs_shutdown = 0
    shifted_by_shutdown_discount = 0
    sample_codes = [normalize_code("211AF1"), normalize_code("332FN1")]
    samples_by_code: dict[str, list[dict[str, Any]]] = {code: [] for code in sample_codes}
    for op in operations:
        by_type[op.work_type_code] += 1
        if op.data_quality_status == "COMPLETE":
            complete += 1
        else:
            missing += 1
        if op.needs_shutdown_date:
            needs_shutdown += 1
        if op.shifted_by_shutdown_discount:
            shifted_by_shutdown_discount += 1
        normalized_equipment = normalize_code(op.equipment_code)
        if normalized_equipment in samples_by_code and len(samples_by_code[normalized_equipment]) < 6:
            samples_by_code[normalized_equipment].append(
                {
                    "equipment_code": op.equipment_code,
                    "work_type": op.work_type_code,
                    "line": op.line_code,
                    "last_date": op.last_date,
                    "previous_scheduled_date": op.previous_scheduled_date,
                    "calendar_due_without_shutdown_discount": op.calendar_due_date,
                    "raw_due_date": op.raw_due_date,
                    "scheduled_date": op.scheduled_date,
                    "execution_condition": op.execution_condition,
                    "shifted_by_shutdown_discount": op.shifted_by_shutdown_discount,
                }
            )
    return {
        "files": len([path for path in ROOT.glob("*.xlsx") if "Master" not in path.name and not path.name.startswith("~$")]),
        "shutdown_windows": len(windows),
        "operations": len(operations),
        "by_type": dict(sorted(by_type.items())),
        "complete": complete,
        "missing_data": missing,
        "needs_shutdown_date": needs_shutdown,
        "shifted_by_shutdown_discount": shifted_by_shutdown_discount,
        "sample_211AF1_332FN1": samples_by_code,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import calculated next maintenance dates from plan files.")
    parser.add_argument("--apply", action="store_true", help="Write calculated plan to the database.")
    parser.add_argument("--as-of", default=datetime.now(timezone(timedelta(hours=3))).date().isoformat())
    args = parser.parse_args()

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
    paths = sorted(path for path in ROOT.glob("*.xlsx") if "Master" not in path.name and not path.name.startswith("~$"))
    windows = collect_shutdown_windows(paths)
    operations = collect_operations(paths, as_of, windows)
    summary = summarize(operations, windows)

    if args.apply:
        database_url = os.environ.get("DATABASE_URL")
        if not database_url:
            raise SystemExit("DATABASE_URL is required for --apply")
        summary["applied"] = apply_to_database(database_url, operations, windows, as_of)

    print(json.dumps(summary, ensure_ascii=False, default=str, indent=2))


if __name__ == "__main__":
    main()
