from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

ROOT = Path(r"f:\lubrication\LUBRICTION PLAN 2026+2027")
NAMESPACE = uuid.UUID("0fd9a65d-f41d-4f4b-b2ce-cf9fc9d0d3e3")


def stable_id(*parts: Any) -> str:
    return str(uuid.uuid5(NAMESPACE, "|".join("" if part is None else str(part) for part in parts)))


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def number(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def quality(required: list[Any]) -> str:
    return "COMPLETE" if all(clean(value) is not None for value in required) else "MISSING_DATA"


def material_code(kind: str, name: str) -> str:
    return f"{kind}-{hashlib.sha1(name.encode('utf-8')).hexdigest()[:10]}"


def execute_values(cur: psycopg2.extensions.cursor, sql: str, rows: list[tuple[Any, ...]]) -> None:
    if rows:
        psycopg2.extras.execute_values(cur, sql, rows, page_size=1000)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--reset", action="store_true", help="Delete existing normalized annual data first")
    args = parser.parse_args()

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL is required")

    areas: dict[str, tuple[Any, ...]] = {}
    lines: dict[str, tuple[Any, ...]] = {}
    materials: dict[str, tuple[Any, ...]] = {}
    equipment: dict[str, tuple[Any, ...]] = {}
    plans: dict[str, tuple[Any, ...]] = {}
    points: list[tuple[Any, ...]] = []
    items: list[tuple[Any, ...]] = []

    for path in sorted(ROOT.glob("* AREA/*.xlsx")):
        area_name = path.parent.name.replace(" AREA", "")
        area_code = area_name.upper().replace(" ", "_")
        area_id = stable_id("area", area_code)
        areas[area_id] = (area_id, area_code, area_name)

        kind = "grease" if "GREASE" in path.name.upper() else "oil"
        work_type_code = "greasing" if kind == "grease" else "oil_change"
        year_match = re.search(r"(20\d{2})", path.name)
        plan_year = int(year_match.group(1)) if year_match else None
        plan_id = stable_id("annual_plan", area_code, plan_year, kind, path.name)
        plans[plan_id] = (plan_id, area_id, plan_year, kind, str(path))

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_index = 1 if wb.worksheets[0].max_row <= 50 and len(wb.worksheets) > 1 else 0
        sheet = wb.worksheets[sheet_index]
        rows = list(sheet.iter_rows(values_only=True))

        for row_number, row in enumerate(rows[3:], start=4):
            if not any(value is not None for value in row):
                continue

            equipment_code = clean(row[0])
            if not equipment_code:
                continue

            equipment_name = clean(row[1])
            part_description = clean(row[2])
            line_code = clean(row[3])
            material_name = clean(row[5] if kind == "grease" else row[4])

            line_id = None
            if line_code:
                line_id = stable_id("line", area_code, line_code)
                lines[line_id] = (line_id, area_id, str(line_code), f"Line {line_code}")

            material_id = None
            if material_name:
                material_id = stable_id("material", kind, material_name)
                materials[material_id] = (
                    material_id,
                    kind,
                    material_code(kind, str(material_name)),
                    str(material_name),
                    "g" if kind == "grease" else "L",
                    "COMPLETE",
                )

            equipment_id = stable_id("equipment", area_code, equipment_code)
            equipment[equipment_id] = (
                equipment_id,
                area_id,
                line_id,
                str(equipment_code),
                str(equipment_name) if equipment_name else None,
                str(part_description) if part_description else None,
                json.dumps({"source_file": str(path), "source_row": row_number}, ensure_ascii=False),
                quality([equipment_code, equipment_name]),
            )

            if kind == "grease":
                point_name = clean(row[4])
                quantity = number(row[8])
                frequency_hours = number(row[7])
                frequency_days = number(row[10])
                last_change_date = clean(row[9])
                last_inspection_date = None
                last_grease_date = clean(row[11])
                quantity_unit = "g"
                execution_condition = "running"
            else:
                point_name = part_description
                quantity = number(row[5])
                frequency_hours = number(row[7])
                frequency_days = number(row[8])
                last_change_date = clean(row[9])
                last_inspection_date = clean(row[10])
                last_grease_date = None
                quantity_unit = "L"
                execution_condition = "shutdown"

            point_id = stable_id("point", str(path), row_number)
            point_status = quality([equipment_code, material_name, quantity])
            original_values = json.dumps(
                {"source_file": str(path), "source_sheet": sheet.title, "source_row": row_number},
                ensure_ascii=False,
            )
            points.append(
                (
                    point_id,
                    equipment_id,
                    work_type_code,
                    material_id,
                    str(point_name) if point_name else None,
                    str(part_description) if part_description else None,
                    execution_condition,
                    quantity,
                    quantity_unit,
                    number(row[6]),
                    frequency_hours,
                    frequency_days,
                    last_change_date if isinstance(last_change_date, date) else None,
                    last_inspection_date if isinstance(last_inspection_date, date) else None,
                    last_grease_date if isinstance(last_grease_date, date) else None,
                    original_values,
                    point_status,
                )
            )
            items.append(
                (
                    stable_id("plan_item", str(path), row_number),
                    plan_id,
                    point_id,
                    quantity,
                    quantity_unit,
                    frequency_hours,
                    frequency_days,
                    original_values,
                    point_status,
                )
            )

        wb.close()

    with psycopg2.connect(database_url, sslmode="require") as conn:
        with conn.cursor() as cur:
            cur.execute("select code, id from maintenance_work_types where code in ('oil_change', 'greasing')")
            work_type_ids = dict(cur.fetchall())
            points = [
                point[:2] + (work_type_ids[point[2]],) + point[3:]
                for point in points
                if point[2] in work_type_ids
            ]

            if args.reset:
                cur.execute(
                    """
                    delete from annual_plan_items;
                    delete from maintenance_points;
                    delete from annual_plans;
                    delete from equipment;
                    delete from materials;
                    delete from production_lines;
                    delete from areas;
                    """
                )

            execute_values(
                cur,
                """
                insert into areas (id, code, name)
                values %s
                on conflict (id) do update set name = excluded.name
                """,
                list(areas.values()),
            )
            execute_values(
                cur,
                """
                insert into production_lines (id, area_id, line_code, name)
                values %s
                on conflict (id) do update set name = excluded.name
                """,
                list(lines.values()),
            )
            execute_values(
                cur,
                """
                insert into materials (id, material_kind, code, name, unit, data_quality_status)
                values %s
                on conflict (id) do update set name = excluded.name, unit = excluded.unit
                """,
                list(materials.values()),
            )
            execute_values(
                cur,
                """
                insert into equipment (
                  id, area_id, production_line_id, equipment_code, name,
                  description, original_values, data_quality_status
                )
                values %s
                on conflict (id) do update set
                  production_line_id = excluded.production_line_id,
                  name = coalesce(excluded.name, equipment.name),
                  description = coalesce(excluded.description, equipment.description),
                  updated_at = now()
                """,
                list(equipment.values()),
            )
            execute_values(
                cur,
                """
                insert into annual_plans (id, area_id, plan_year, material_kind, source_file)
                values %s
                on conflict (id) do update set updated_at = now()
                """,
                list(plans.values()),
            )
            execute_values(
                cur,
                """
                insert into maintenance_points (
                  id, equipment_id, work_type_id, material_id, point_name,
                  part_description, execution_condition, quantity, quantity_unit,
                  running_hours_per_day, frequency_hours, frequency_days,
                  last_change_date, last_inspection_date, last_grease_date,
                  original_values, data_quality_status
                )
                values %s
                on conflict (id) do update set updated_at = now()
                """,
                points,
            )
            execute_values(
                cur,
                """
                insert into annual_plan_items (
                  id, annual_plan_id, maintenance_point_id, planned_quantity,
                  planned_quantity_unit, frequency_hours, frequency_days,
                  original_values, data_quality_status
                )
                values %s
                on conflict (id) do update set updated_at = now()
                """,
                items,
            )
            conn.commit()

    print(
        {
            "areas": len(areas),
            "production_lines": len(lines),
            "materials": len(materials),
            "equipment": len(equipment),
            "maintenance_points": len(points),
            "annual_plans": len(plans),
            "annual_plan_items": len(items),
        }
    )


if __name__ == "__main__":
    main()
