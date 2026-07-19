from __future__ import annotations

import argparse
import json
import os
import re
import uuid
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

ROOT = Path(r"f:\lubrication")
NAMESPACE = uuid.UUID("e1c61183-7483-48fa-82f2-fd45cfaf17c1")


def stable_id(*parts: Any) -> str:
    return str(uuid.uuid5(NAMESPACE, "|".join("" if part is None else str(part) for part in parts)))


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\n", " ").strip()


def canonical_code(value: Any) -> str:
    code = clean(value).upper()
    code = re.sub(r"\s+", "", code)
    code = code.replace("-", "").replace("/", "")
    code = re.sub(r"([A-Z]+)0+([0-9])", r"\1\2", code)
    return code


def equipment_code_like(value: Any) -> bool:
    code = clean(value).upper()
    return bool(re.match(r"^[0-9]{3}[A-Z]{2}[A-Z0-9 /-]*$", code))


def master_path() -> Path:
    matches = list(ROOT.glob("Factory Maintenance System*Master*.xlsx"))
    if not matches:
        raise FileNotFoundError("Master workbook was not found")
    return matches[0]


def collect_master_equipment(path: Path) -> dict[str, dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    equipment: dict[str, dict[str, Any]] = {}
    source_sheets = ["oIL-Master-Line", "Grease-Master-Line"]

    for sheet_name in source_sheets:
        sheet = workbook[sheet_name]
        headers = [clean(value) or f"col_{index + 1}" for index, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))]
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_code = clean(row[0] if len(row) > 0 else None).upper()
            if not equipment_code_like(raw_code):
                continue

            code_key = canonical_code(raw_code)
            name = clean(row[1] if len(row) > 1 else None)
            line = clean(row[2] if len(row) > 2 else None) or "بدون مكان"
            description = clean(row[3] if len(row) > 3 else None)

            if code_key not in equipment:
                equipment[code_key] = {
                    "equipment_code": raw_code,
                    "name": name or raw_code,
                    "line": line,
                    "descriptions": [],
                    "source_rows": [],
                    "master_rows": [],
                }

            if name and not equipment[code_key]["name"]:
                equipment[code_key]["name"] = name
            if description and description not in equipment[code_key]["descriptions"]:
                equipment[code_key]["descriptions"].append(description)
            equipment[code_key]["source_rows"].append({"sheet": sheet_name, "row": row_number})
            equipment[code_key]["master_rows"].append(
                {
                    "sheet": sheet_name,
                    "row": row_number,
                    "columns": {
                        str(headers[index] if index < len(headers) else f"col_{index + 1}"): clean(value)
                        for index, value in enumerate(row)
                        if clean(value)
                    },
                }
            )

    workbook.close()
    return equipment


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync equipment list from the master workbook.")
    parser.add_argument("--apply", action="store_true", help="Write changes to the database.")
    args = parser.parse_args()

    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL is required")

    path = master_path()
    master_equipment = collect_master_equipment(path)
    master_keys = set(master_equipment)

    stats = {
        "master_file": str(path),
        "master_equipment": len(master_equipment),
        "updated": 0,
        "inserted": 0,
        "hidden_not_in_master": 0,
    }

    with psycopg2.connect(database_url, sslmode="require") as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                insert into areas (id, code, name)
                values (%s, 'MASTER', 'المعدات الرئيسية')
                on conflict (code) do update set name = excluded.name
                returning id
                """,
                (stable_id("area", "MASTER"),),
            )
            master_area_id = cur.fetchone()["id"]

            cur.execute("select id, equipment_code, original_values from equipment")
            rows = cur.fetchall()
            by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in rows:
                if (row.get("original_values") or {}).get("source_mode") == "calculated_plan_equipment":
                    continue
                by_key[canonical_code(row["equipment_code"])].append(row)

            if args.apply:
                for key, item in master_equipment.items():
                    original_values = {
                        "source_mode": "master_equipment",
                        "master_file": str(path),
                        "master_line": item["line"],
                        "master_source_rows": item["source_rows"],
                        "master_rows": item["master_rows"],
                        "in_master": True,
                    }
                    description = " / ".join(item["descriptions"][:6]) or None
                    matches = by_key.get(key, [])

                    if matches:
                        preferred_id = matches[0]["id"]
                        for match in matches:
                            is_preferred = match["id"] == preferred_id
                            duplicate_values = {} if is_preferred else {"duplicate_master_equipment": True}
                            cur.execute(
                                """
                                update equipment
                                set name = %s,
                                    description = coalesce(%s, description),
                                    original_values = coalesce(original_values, '{}'::jsonb) || %s::jsonb,
                                    data_quality_status = 'COMPLETE',
                                    is_active = %s,
                                    updated_at = now()
                                where id = %s
                                """,
                                (
                                    item["name"],
                                    description,
                                    json.dumps({**original_values, **duplicate_values}, ensure_ascii=False),
                                    is_preferred,
                                    match["id"],
                                ),
                            )
                            stats["updated"] += 1
                    else:
                        cur.execute(
                            """
                            insert into equipment (
                              id, area_id, equipment_code, name, description,
                              original_values, data_quality_status, is_active
                            )
                            values (%s, %s, %s, %s, %s, %s::jsonb, 'COMPLETE', true)
                            """,
                            (
                                stable_id("master-equipment", key),
                                master_area_id,
                                item["equipment_code"],
                                item["name"],
                                description,
                                json.dumps(original_values, ensure_ascii=False),
                            ),
                        )
                        stats["inserted"] += 1

                for key, matches in by_key.items():
                    if key in master_keys:
                        continue
                    for match in matches:
                        cur.execute(
                            """
                            update equipment
                            set is_active = false,
                                original_values = coalesce(original_values, '{}'::jsonb) || '{"in_master": false}'::jsonb,
                                updated_at = now()
                            where id = %s
                            """,
                            (match["id"],),
                        )
                        stats["hidden_not_in_master"] += 1

                conn.commit()
            else:
                matched_existing = sum(len(by_key.get(key, [])) for key in master_keys)
                stats["updated"] = matched_existing
                stats["inserted"] = len([key for key in master_keys if key not in by_key])
                stats["hidden_not_in_master"] = sum(len(matches) for key, matches in by_key.items() if key not in master_keys)

    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
