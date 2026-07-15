from __future__ import annotations

import hashlib
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2

ROOT = Path(r"f:\lubrication\LUBRICTION PLAN 2026+2027")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def as_number(value: Any) -> float | None:
    value = clean(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def material_code(kind: str, name: str) -> str:
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:10]
    return f"{kind}-{digest}"


def quality(required: list[Any]) -> str:
    return "COMPLETE" if all(clean(value) is not None for value in required) else "MISSING_DATA"


def one(cur: psycopg2.extensions.cursor, query: str, params: tuple[Any, ...]) -> str:
    cur.execute(query, params)
    return cur.fetchone()[0]


def main() -> None:
    database_url = os.environ.get("DATABASE_URL")
    if not database_url:
        raise SystemExit("DATABASE_URL is required")

    with psycopg2.connect(database_url, sslmode="require") as conn:
        with conn.cursor() as cur:
            work_type_ids = {}
            for code in ["oil_change", "greasing"]:
                cur.execute("select id from maintenance_work_types where code = %s", (code,))
                work_type_ids[code] = cur.fetchone()[0]

            inserted = {"areas": 0, "plans": 0, "items": 0, "points": 0}

            for path in sorted(ROOT.glob("* AREA/*.xlsx")):
                area_name = path.parent.name.replace(" AREA", "")
                area_code = area_name.upper().replace(" ", "_")
                kind = "grease" if "GREASE" in path.name.upper() else "oil"
                work_type_id = work_type_ids["greasing" if kind == "grease" else "oil_change"]
                year_match = re.search(r"(20\d{2})", path.name)
                plan_year = int(year_match.group(1)) if year_match else None

                area_id = one(
                    cur,
                    """
                    insert into areas (code, name)
                    values (%s, %s)
                    on conflict (code) do update set name = excluded.name
                    returning id
                    """,
                    (area_code, area_name),
                )

                plan_id = one(
                    cur,
                    """
                    insert into annual_plans (area_id, plan_year, material_kind, source_file)
                    values (%s, %s, %s, %s)
                    on conflict (area_id, plan_year, material_kind, source_file)
                    do update set updated_at = now()
                    returning id
                    """,
                    (area_id, plan_year, kind, str(path)),
                )

                cur.execute("select id from import_files where source_path = %s order by created_at desc limit 1", (str(path),))
                import_file = cur.fetchone()
                import_file_id = import_file[0] if import_file else None

                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                sheet = wb.worksheets[1] if wb.worksheets[0].max_row <= 50 and len(wb.worksheets) > 1 else wb.worksheets[0]
                rows = list(sheet.iter_rows(values_only=True))
                headers = [clean(value) or f"col_{index + 1}" for index, value in enumerate(rows[2])]

                for row_number, row in enumerate(rows[3:], start=4):
                    if not any(value is not None for value in row):
                        continue

                    values = {headers[index]: clean(value) for index, value in enumerate(row)}
                    equipment_code = clean(row[0])
                    equipment_name = clean(row[1])
                    part_description = clean(row[2])
                    line_code = clean(row[3])
                    material_name = clean(row[5] if kind == "grease" else row[4])

                    if not equipment_code:
                        continue

                    production_line_id = None
                    if line_code:
                        production_line_id = one(
                            cur,
                            """
                            insert into production_lines (area_id, line_code, name)
                            values (%s, %s, %s)
                            on conflict (area_id, line_code) do update set name = excluded.name
                            returning id
                            """,
                            (area_id, str(line_code), f"Line {line_code}"),
                        )

                    material_id = None
                    if material_name:
                        code = material_code(kind, str(material_name))
                        material_id = one(
                            cur,
                            """
                            insert into materials (material_kind, code, name, unit, data_quality_status)
                            values (%s, %s, %s, %s, 'COMPLETE')
                            on conflict (material_kind, code) do update set name = excluded.name
                            returning id
                            """,
                            (kind, code, str(material_name), "g" if kind == "grease" else "L"),
                        )

                    equipment_status = quality([equipment_code, equipment_name])
                    equipment_id = one(
                        cur,
                        """
                        insert into equipment (
                          area_id, production_line_id, equipment_code, name, description,
                          original_values, data_quality_status
                        )
                        values (%s, %s, %s, %s, %s, %s::jsonb, %s)
                        on conflict (area_id, equipment_code)
                        do update set
                          name = coalesce(excluded.name, equipment.name),
                          production_line_id = coalesce(excluded.production_line_id, equipment.production_line_id),
                          updated_at = now()
                        returning id
                        """,
                        (
                            area_id,
                            production_line_id,
                            str(equipment_code),
                            str(equipment_name) if equipment_name else None,
                            str(part_description) if part_description else None,
                            "{}",
                            equipment_status,
                        ),
                    )

                    source_row_id = None
                    if import_file_id:
                        cur.execute(
                            """
                            select id
                            from imported_rows
                            where import_file_id = %s
                              and sheet_index = %s
                              and row_number = %s
                            limit 1
                            """,
                            (import_file_id, 1 if wb.worksheets[0].max_row <= 50 and len(wb.worksheets) > 1 else 0, row_number),
                        )
                        source_row = cur.fetchone()
                        source_row_id = source_row[0] if source_row else None

                    if kind == "grease":
                        point_name = clean(row[4])
                        quantity = as_number(row[8])
                        frequency_hours = as_number(row[7])
                        frequency_days = as_number(row[10])
                        last_change_date = clean(row[9])
                        last_grease_date = clean(row[11])
                        quantity_unit = "g"
                        execution_condition = "running"
                    else:
                        point_name = part_description
                        quantity = as_number(row[5])
                        frequency_hours = as_number(row[7])
                        frequency_days = as_number(row[8])
                        last_change_date = clean(row[9])
                        last_grease_date = None
                        quantity_unit = "L"
                        execution_condition = "shutdown"

                    point_status = quality([equipment_code, material_name, quantity])
                    point_id = one(
                        cur,
                        """
                        insert into maintenance_points (
                          equipment_id, work_type_id, material_id, point_name, part_description,
                          execution_condition, quantity, quantity_unit, running_hours_per_day,
                          frequency_hours, frequency_days, last_change_date, last_inspection_date,
                          last_grease_date, original_values, data_quality_status
                        )
                        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                        returning id
                        """,
                        (
                            equipment_id,
                            work_type_id,
                            material_id,
                            str(point_name) if point_name else None,
                            str(part_description) if part_description else None,
                            execution_condition,
                            quantity,
                            quantity_unit,
                            as_number(row[6]),
                            frequency_hours,
                            frequency_days,
                            last_change_date if isinstance(last_change_date, date) else None,
                            clean(row[10]) if kind == "oil" and isinstance(clean(row[10]), date) else None,
                            last_grease_date if isinstance(last_grease_date, date) else None,
                            "{}",
                            point_status,
                        ),
                    )

                    cur.execute(
                        """
                        insert into annual_plan_items (
                          annual_plan_id, maintenance_point_id, source_row_id, planned_quantity,
                          planned_quantity_unit, frequency_hours, frequency_days, original_values,
                          data_quality_status
                        )
                        values (%s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                        """,
                        (
                            plan_id,
                            point_id,
                            source_row_id,
                            quantity,
                            quantity_unit,
                            frequency_hours,
                            frequency_days,
                            "{}",
                            point_status,
                        ),
                    )
                    inserted["items"] += 1
                    inserted["points"] += 1

                wb.close()
                inserted["plans"] += 1
                conn.commit()

            print(inserted)


if __name__ == "__main__":
    main()
